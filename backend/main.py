"""Exhibitor Extraction Web App - Backend

FastAPI backend for processing trade show exhibitor lists, enriching contacts via Seamless.ai API,
and outputting XLSX files to Google Drive.
"""

import asyncio
import json
import os
import re
import sys
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
import openpyxl
from fastapi import BackgroundTasks, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google.oauth2.service_account import Credentials as ServiceCredentials
from googleapiclient.discovery import build
from pydantic import BaseModel

# Load environment variables
from dotenv import load_dotenv
load_dotenv(Path.home() / ".openclaw" / ".env")

# Add floorplan module to path (relative for deployment)
FLOORPLAN_PATH = Path(__file__).parent / "floorplan"
sys.path.insert(0, str(FLOORPLAN_PATH))

# Constants
SHEET_ID = "1Yhamt9si8Hs64g0q4JVwlEyghXsgDUXT"
DRIVE_FOLDER_ID = "17xvwirTNTfH5HmKmXIY7JYW_XoxJ1T-5"
# Template path (relative for deployment)
TEMPLATE_PATH = Path(__file__).parent / "AE_ShowList_Template.xlsx"

# Title search tiers for Seamless.ai
TIER_1 = [
    "Event Director", "Event Manager", "Event Coordinator",
    "Trade Show Director", "Trade Show Manager", "Trade Show Coordinator",
    "Field Event Director", "Field Event Manager", "Field Event Coordinator"
]

TIER_2 = [
    "Marketing Director", "Marketing Manager", "Marketing Coordinator",
    "Business Development Director", "Business Development Manager",
    "Sales Director", "Sales Manager", "EMEA Director", "EMEA Manager",
    "US Director", "US Manager"
]

TIER_3 = ["Founder", "Co-Founder", "President", "CEO", "VP Sales", "VP Revenue"]

EXCLUDE_TITLES = [
    "intern", "ops", "operations", "assistant", "legal", "counsel",
    "IT", "engineer", "developer", "HR", "recruiting", "finance",
    "accounting", "admin", "customer success", "support"
]

# Storage for job status
jobs: Dict[str, Dict[str, Any]] = {}

app = FastAPI(
    title="Exhibitor Extraction API",
    description="Process trade show exhibitor lists with contact enrichment",
    version="1.0.0"
)

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://exhibitor-extraction-ui.onrender.com",
        "http://localhost:5173",  # Local development
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================================
# Models
# ============================================================================

class PendingRequest(BaseModel):
    row_index: int
    date_requested: Optional[str] = None
    show_name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    location: Optional[str] = None
    floorplan_url: Optional[str] = None
    exhibitor_list_url: Optional[str] = None
    delivered: Optional[str] = None
    exhibitors: Optional[int] = None
    large_booths: Optional[int] = None
    missing_contact_info: Optional[int] = None
    list_cost: Optional[float] = None


class ProcessRequest(BaseModel):
    row_indices: List[int]


class JobStatus(BaseModel):
    job_id: str
    status: str  # "pending", "processing", "completed", "failed"
    progress: int  # 0-100
    message: str
    results: Optional[List[Dict[str, Any]]] = None
    error: Optional[str] = None


# ============================================================================
# Google API Helpers
# ============================================================================

def get_google_credentials():
    """Get Google API credentials from environment, service account file, or OAuth token."""
    
    # Method 1: Service account from environment variables
    service_account_email = os.getenv("GOOGLE_SERVICE_ACCOUNT_EMAIL")
    private_key = os.getenv("GOOGLE_PRIVATE_KEY")
    
    if service_account_email and private_key:
        # Handle escaped newlines in private key
        private_key = private_key.replace("\\n", "\n")
        try:
            credentials = ServiceCredentials.from_service_account_info({
                "type": "service_account",
                "client_email": service_account_email,
                "private_key": private_key,
                "token_uri": "https://oauth2.googleapis.com/token"
            }, scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive.file"
            ])
            return credentials
        except Exception as e:
            print(f"Failed to create service account credentials: {e}")
    
    # Method 2: Service account JSON file
    service_account_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    if not service_account_path:
        # Check common locations
        possible_paths = [
            Path.home() / ".openclaw" / "google" / "service-account.json",
            Path.home() / ".openclaw" / "google" / "credentials.json",
            Path("/app/credentials/service-account.json"),  # Render deployment
        ]
        for path in possible_paths:
            if path.exists():
                service_account_path = str(path)
                break
    
    if service_account_path and Path(service_account_path).exists():
        try:
            credentials = ServiceCredentials.from_service_account_file(
                service_account_path,
                scopes=[
                    "https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive.file"
                ]
            )
            return credentials
        except Exception as e:
            print(f"Failed to load service account file: {e}")
    
    # Method 3: OAuth2 token file (for local development)
    token_path = Path.home() / ".openclaw" / "google" / "token.json"
    if token_path.exists():
        try:
            creds = Credentials.from_authorized_user_file(
                str(token_path),
                ["https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive.file"]
            )
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
                # Save refreshed token
                token_path.write_text(creds.to_json())
            if creds and creds.valid:
                return creds
        except Exception as e:
            print(f"Failed to load OAuth credentials: {e}")
    
    # No valid credentials
    raise HTTPException(
        status_code=500,
        detail="No valid Google credentials found. Please configure GOOGLE_SERVICE_ACCOUNT_EMAIL and GOOGLE_PRIVATE_KEY, or provide a service-account.json file."
    )


def get_sheets_service():
    """Get Google Sheets service."""
    creds = get_google_credentials()
    return build("sheets", "v4", credentials=creds)


def get_drive_service():
    """Get Google Drive service."""
    creds = get_google_credentials()
    return build("drive", "v3", credentials=creds)


# ============================================================================
# Seamless.ai API Helpers
# ============================================================================

async def search_seamless_contacts(
    client: httpx.AsyncClient,
    company_names: List[str],
    titles: List[str],
    api_key: str
) -> List[Dict[str, Any]]:
    """Search for contacts via Seamless.ai API."""
    
    # Batch up to 100 companies, max 10 titles per request
    all_contacts = []
    
    for i in range(0, len(company_names), 100):
        batch = company_names[i:i+100]
        
        payload = {
            "companyName": batch,
            "jobTitle": titles[:10],
            "department": ["Sales", "Marketing"],
            "limit": 50
        }
        
        headers = {
            "Token": api_key,
            "Content-Type": "application/json"
        }
        
        try:
            response = await client.post(
                "https://api.seamless.ai/api/client/v1/contacts/search",
                json=payload,
                headers=headers,
                timeout=60.0
            )
            
            if response.status_code == 200:
                data = response.json()
                contacts = data.get("contacts", []) or data.get("results", []) or data.get("data", []) or []
                all_contacts.extend(contacts)
            elif response.status_code == 429:
                # Rate limited - wait and retry
                await asyncio.sleep(60)
                response = await client.post(
                    "https://api.seamless.ai/api/client/v1/contacts/search",
                    json=payload,
                    headers=headers,
                    timeout=60.0
                )
                if response.status_code == 200:
                    data = response.json()
                    contacts = data.get("contacts", []) or data.get("results", []) or data.get("data", []) or []
                    all_contacts.extend(contacts)
            else:
                print(f"Seamless.ai API error: {response.status_code} - {response.text[:200]}")
            
            # Rate limiting: 60 requests/minute (wait 1.1 seconds between requests)
            await asyncio.sleep(1.1)
            
        except Exception as e:
            print(f"Seamless.ai API error: {e}")
    
    return all_contacts


def filter_contacts(contacts: List[Dict[str, Any]], max_per_company: int = 5) -> Dict[str, List[Dict[str, Any]]]:
    """Filter and deduplicate contacts by company."""
    
    by_company: Dict[str, List[Dict[str, Any]]] = {}
    seen_emails = set()
    
    for contact in contacts:
        # Normalize email
        email = (contact.get("email") or contact.get("Email") or "").lower().strip()
        if not email or email in seen_emails:
            continue
        seen_emails.add(email)
        
        # Filter by excluded titles
        title = (contact.get("title") or contact.get("jobTitle") or contact.get("Title") or "").lower()
        if any(ex.lower() in title for ex in EXCLUDE_TITLES):
            continue
        
        # Get company name
        company = (contact.get("company") or contact.get("companyName") or contact.get("Company") or "").strip()
        if not company:
            continue
        
        # Normalize company name for matching
        company_key = company.lower()
        
        if company_key not in by_company:
            by_company[company_key] = []
        
        if len(by_company[company_key]) < max_per_company:
            by_company[company_key].append(contact)
    
    return by_company


async def enrich_exhibitors(
    client: httpx.AsyncClient,
    exhibitors: List[Dict[str, Any]],
    api_key: str
) -> List[Dict[str, Any]]:
    """Enrich exhibitor list with contact information from Seamless.ai."""
    
    if not api_key:
        print("No Seamless.ai API key provided, skipping enrichment")
        return exhibitors
    
    # Get unique company names
    company_names = list(set(
        ex["company_name"] for ex in exhibitors 
        if ex.get("company_name")
    ))
    
    if not company_names:
        return exhibitors
    
    # Search with tiered titles
    all_contacts = []
    
    # Tier 1 search
    print(f"Searching Tier 1 titles for {len(company_names)} companies...")
    contacts = await search_seamless_contacts(client, company_names, TIER_1, api_key)
    all_contacts.extend(contacts)
    
    # Check which companies need more contacts
    by_company = filter_contacts(all_contacts)
    needs_more = [c for c in company_names if len(by_company.get(c.lower(), [])) < 3]
    
    # Tier 2 search for companies with <3 contacts
    if needs_more:
        print(f"Searching Tier 2 titles for {len(needs_more)} companies needing more contacts...")
        contacts = await search_seamless_contacts(client, needs_more, TIER_2, api_key)
        all_contacts.extend(contacts)
        
        by_company = filter_contacts(all_contacts)
        needs_more = [c for c in company_names if len(by_company.get(c.lower(), [])) < 3]
    
    # Tier 3 search for companies still with <3 contacts
    if needs_more:
        print(f"Searching Tier 3 titles for {len(needs_more)} companies needing more contacts...")
        contacts = await search_seamless_contacts(client, needs_more, TIER_3, api_key)
        all_contacts.extend(contacts)
    
    # Final filter
    by_company = filter_contacts(all_contacts)
    print(f"Found contacts for {len(by_company)} companies")
    
    # Merge contacts into exhibitor records
    enriched = []
    for ex in exhibitors:
        company = ex.get("company_name", "")
        contacts = by_company.get(company.lower(), [])
        
        if contacts:
            for contact in contacts[:5]:  # Max 5 contacts per company
                enriched.append({
                    **ex,
                    "contact_first_name": contact.get("firstName") or contact.get("first_name") or contact.get("FirstName", ""),
                    "contact_last_name": contact.get("lastName") or contact.get("last_name") or contact.get("LastName", ""),
                    "contact_title": contact.get("title") or contact.get("jobTitle") or contact.get("Title", ""),
                    "contact_email": contact.get("email") or contact.get("Email", ""),
                    "contact_linkedin": contact.get("linkedin") or contact.get("linkedinUrl") or contact.get("Linkedin", ""),
                    "company_phone": contact.get("phone") or contact.get("companyPhone") or contact.get("Phone", ""),
                    "company_website": contact.get("website") or contact.get("companyWebsite") or contact.get("Website", ""),
                    "company_linkedin": contact.get("companyLinkedin") or contact.get("CompanyLinkedin", ""),
                    "address_line1": contact.get("address") or contact.get("addressLine1") or contact.get("Address", ""),
                    "city": contact.get("city") or contact.get("addressTownCity") or contact.get("City", ""),
                    "state": contact.get("state") or contact.get("addressCountyState") or contact.get("State", ""),
                    "country": contact.get("country") or contact.get("addressCountry") or contact.get("Country", ""),
                    "postal_code": contact.get("postalCode") or contact.get("addressZipPostCode") or contact.get("PostalCode", "")
                })
        else:
            # No contacts found, keep original exhibitor record
            enriched.append({
                **ex,
                "contact_first_name": "",
                "contact_last_name": "",
                "contact_title": "",
                "contact_email": "",
                "contact_linkedin": "",
                "company_phone": "",
                "company_website": "",
                "company_linkedin": "",
                "address_line1": "",
                "city": "",
                "state": "",
                "country": "",
                "postal_code": ""
            })
    
    return enriched


# ============================================================================
# Exhibitor Extraction
# ============================================================================

def extract_year_from_url(url: str) -> Optional[int]:
    """Extract year from URL patterns."""
    if not url:
        return None
    
    # Pattern: .../2026/... in path
    match = re.search(r'/(\d{4})/', url)
    if match:
        return int(match.group(1))
    
    # Pattern: ..._2026_... in filename
    match = re.search(r'_(\d{4})_', url)
    if match:
        return int(match.group(1))
    
    # Pattern: ...26... in URL (e.g., aaos26.mapyourshow.com -> 2026)
    match = re.search(r'(\d{2})\.', url)
    if match:
        year_suffix = int(match.group(1))
        current_century = datetime.now().year // 100 * 100
        year = current_century + year_suffix
        # Sanity check: year should be within reasonable range
        if 2024 <= year <= 2030:
            return year
    
    return None


async def extract_exhibitors(floorplan_url: str, exhibitor_list_url: Optional[str] = None) -> List[Dict[str, Any]]:
    """Extract exhibitors from floorplan or exhibitor list URL."""
    
    exhibitors = []
    
    # Priority: HTML exhibitor list → PDF floorplan → Image floorplan
    if exhibitor_list_url:
        try:
            print(f"Extracting from exhibitor list: {exhibitor_list_url}")
            
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    exhibitor_list_url,
                    headers={'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36'},
                    timeout=30.0
                )
                response.raise_for_status()
                html = response.text
            
            # Use the exhibitor_list_extract module
            from exhibitor_list_extract import extract_from_html_text, deduplicate_records
            
            records = extract_from_html_text(html, exhibitor_list_url)
            records = deduplicate_records(records)
            
            exhibitors = [
                {
                    "booth_number": r.booth_number,
                    "company_name": r.exhibitor_name,
                    "booth_width": None,
                    "booth_height": None
                }
                for r in records
                if r.confidence >= 0.40
            ]
            
            if exhibitors:
                print(f"Extracted {len(exhibitors)} exhibitors from exhibitor list")
                return exhibitors
                
        except Exception as e:
            print(f"Exhibitor list extraction failed: {e}")
    
    # Fall back to floorplan extraction
    if floorplan_url:
        try:
            print(f"Extracting from floorplan: {floorplan_url}")
            
            # Check if PDF
            if floorplan_url.lower().endswith('.pdf'):
                try:
                    import pymupdf
                except ImportError:
                    import fit as pymupdf
                
                # Download PDF
                async with httpx.AsyncClient() as client:
                    response = await client.get(floorplan_url, timeout=60.0)
                    response.raise_for_status()
                
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                    tmp.write(response.content)
                    tmp_path = tmp.name
                
                try:
                    doc = pymupdf.open(tmp_path)
                    for page_num in range(len(doc)):
                        page = doc[page_num]
                        text = page.get_text()
                        
                        # Extract booth patterns from text
                        booth_pattern = r"[A-Z]?\d{1,4}[A-Z]?"
                        lines = text.split('\n')
                        
                        for line in lines:
                            booths = re.findall(booth_pattern, line.upper())
                            company = re.sub(booth_pattern, '', line).strip(' ,;-|')
                            
                            if booths and company and len(company) > 3:
                                for booth in booths:
                                    exhibitors.append({
                                        "booth_number": booth,
                                        "company_name": company,
                                        "booth_width": None,
                                        "booth_height": None
                                    })
                    
                    doc.close()
                finally:
                    os.unlink(tmp_path)
            
            else:
                # Image floorplan - use floorplan_extract module
                from floorplan_extract import load_image, detect_grid_candidates, ocr_regions
                
                # Download image
                async with httpx.AsyncClient() as client:
                    response = await client.get(floorplan_url, timeout=60.0)
                    response.raise_for_status()
                
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                    tmp.write(response.content)
                    tmp_path = tmp.name
                
                try:
                    img = load_image(path=tmp_path)
                    regions = detect_grid_candidates(img)
                    ocr_results = ocr_regions(img, regions)
                    
                    for ocr in ocr_results:
                        if ocr.booth_numbers:
                            for booth in ocr.booth_numbers.split():
                                exhibitors.append({
                                    "booth_number": booth,
                                    "company_name": ocr.organisation_name,
                                    "booth_width": None,
                                    "booth_height": None
                                })
                finally:
                    os.unlink(tmp_path)
            
            print(f"Extracted {len(exhibitors)} exhibitors from floorplan")
        
        except Exception as e:
            print(f"Floorplan extraction failed: {e}")
    
    return exhibitors


# ============================================================================
# XLSX Generation
# ============================================================================

def generate_xlsx(exhibitors: List[Dict[str, Any]], output_path: str, show_name: str = "") -> str:
    """Generate XLSX file from exhibitor data using template."""
    
    # Load template
    if TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        # Find the last row with data or use row 2
        start_row = 2
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exhibitors"
        # Headers
        headers = [
            "Booth", "Company", "Linkedin_Company", "Phone|Company", "Website",
            "BoothWidth", "BoothLength", "Name_First", "Name_Last", "Position",
            "Email|Contact", "Linkedin_Contact", "Address_Line1", "Address_TownCity",
            "Address_CountyState", "Address_Country", "Address_ZipPostCode"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        start_row = 2
    
    # Add exhibitor data
    for ex in exhibitors:
        ws.cell(row=start_row, column=1, value=ex.get("booth_number", ""))
        ws.cell(row=start_row, column=2, value=ex.get("company_name", ""))
        ws.cell(row=start_row, column=3, value=ex.get("company_linkedin", ""))
        ws.cell(row=start_row, column=4, value=ex.get("company_phone", ""))
        ws.cell(row=start_row, column=5, value=ex.get("company_website", ""))
        ws.cell(row=start_row, column=6, value=ex.get("booth_width"))
        ws.cell(row=start_row, column=7, value=ex.get("booth_height"))
        ws.cell(row=start_row, column=8, value=ex.get("contact_first_name", ""))
        ws.cell(row=start_row, column=9, value=ex.get("contact_last_name", ""))
        ws.cell(row=start_row, column=10, value=ex.get("contact_title", ""))
        ws.cell(row=start_row, column=11, value=ex.get("contact_email", ""))
        ws.cell(row=start_row, column=12, value=ex.get("contact_linkedin", ""))
        ws.cell(row=start_row, column=13, value=ex.get("address_line1", ""))
        ws.cell(row=start_row, column=14, value=ex.get("city", ""))
        ws.cell(row=start_row, column=15, value=ex.get("state", ""))
        ws.cell(row=start_row, column=16, value=ex.get("country", ""))
        ws.cell(row=start_row, column=17, value=ex.get("postal_code", ""))
        start_row += 1
    
    wb.save(output_path)
    return output_path


def upload_to_drive(file_path: str, filename: str) -> str:
    """Upload file to Google Drive and return shareable link."""
    
    drive_service = get_drive_service()
    
    # Upload file
    from googleapiclient.http import MediaFileUpload
    
    file_metadata = {
        "name": filename,
        "parents": [DRIVE_FOLDER_ID]
    }
    
    media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    file = drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id"
    ).execute()
    
    file_id = file.get("id")
    
    # Make file shareable
    drive_service.permissions().create(
        fileId=file_id,
        body={"type": "anyone", "role": "reader"},
        fields="id"
    ).execute()
    
    return f"https://drive.google.com/file/d/{file_id}/view"


# ============================================================================
# Processing Pipeline
# ============================================================================

async def process_request(
    job_id: str,
    row_indices: List[int],
    seamless_api_key: str
):
    """Process selected requests in background."""
    
    try:
        jobs[job_id]["status"] = "processing"
        jobs[job_id]["message"] = "Reading pending requests from Google Sheet..."
        jobs[job_id]["progress"] = 5
        
        # Get pending requests
        sheets_service = get_sheets_service()
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range="A:L"
        ).execute()
        
        rows = result.get("values", [])
        if len(rows) < 2:
            jobs[job_id]["status"] = "failed"
            jobs[job_id]["error"] = "No data found in spreadsheet"
            return
        
        # Parse header
        headers = rows[0]
        header_map = {}
        for i, h in enumerate(headers):
            key = h.lower().replace(" ", "_") if h else f"col_{i}"
            header_map[key] = i
        
        results = []
        total = len(row_indices)
        
        async with httpx.AsyncClient() as client:
            for idx, row_idx in enumerate(row_indices):
                if row_idx >= len(rows):
                    continue
                
                row = rows[row_idx]
                
                # Helper to safely get row value
                def get_row_val(key: str, default: str = "") -> str:
                    col_idx = header_map.get(key, -1)
                    if col_idx >= 0 and col_idx < len(row):
                        return row[col_idx]
                    return default
                
                # Parse row data
                show_name = get_row_val("show_name")
                location = get_row_val("location")
                floorplan_url = get_row_val("floorplan")
                exhibitor_list_url = get_row_val("exhibitor_list")
                start_date = get_row_val("start")
                
                jobs[job_id]["message"] = f"Extracting exhibitors for: {show_name}"
                jobs[job_id]["progress"] = 10 + int((idx / total) * 40)
                
                # Extract exhibitors
                exhibitors = await extract_exhibitors(floorplan_url, exhibitor_list_url)
                
                if not exhibitors:
                    results.append({
                        "show_name": show_name,
                        "status": "failed",
                        "error": "No exhibitors extracted"
                    })
                    continue
                
                jobs[job_id]["message"] = f"Enriching contacts for: {show_name}"
                jobs[job_id]["progress"] = 50 + int((idx / total) * 30)
                
                # Enrich with Seamless.ai
                enriched = await enrich_exhibitors(client, exhibitors, seamless_api_key)
                
                jobs[job_id]["message"] = f"Generating XLSX for: {show_name}"
                jobs[job_id]["progress"] = 80 + int((idx / total) * 10)
                
                # Generate filename
                year = extract_year_from_url(floorplan_url) or datetime.now().year
                
                # Parse start date for YYYY_MM
                yyyymm = datetime.now().strftime("%Y_%m")
                if start_date:
                    try:
                        # Handle MM/DD/YY format
                        dt = datetime.strptime(start_date, "%m/%d/%y")
                        yyyymm = dt.strftime("%Y_%m")
                        if not year:
                            year = dt.year
                    except:
                        pass
                
                # Format location
                city_state = ""
                if location:
                    parts = location.split(",")
                    city_state = parts[0].strip() if parts else location
                    if len(parts) > 1:
                        state = parts[1].strip()[:2]
                        city_state += f", {state}"
                
                # Clean filename
                safe_show_name = re.sub(r'[<>:"/\\|?*]', '', show_name)
                filename = f"{yyyymm}_{safe_show_name} {year}, {city_state}.xlsx"
                
                # Generate XLSX
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp_path = tmp.name
                
                try:
                    generate_xlsx(enriched, tmp_path, show_name)
                    
                    # Upload to Google Drive
                    jobs[job_id]["message"] = f"Uploading to Google Drive: {show_name}"
                    drive_link = upload_to_drive(tmp_path, filename)
                    
                    # Update sheet with counts
                    exhibitor_count = len(set(e.get("company_name") for e in enriched if e.get("company_name")))
                    large_booths = sum(
                        1 for e in enriched 
                        if (e.get("booth_width") or 0) >= 20 and (e.get("booth_height") or 0) >= 20
                    )
                    
                    # Update the tracking sheet (column I = exhibitors, J = 20x20+, K = missing contact)
                    sheets_service.spreadsheets().values().update(
                        spreadsheetId=SHEET_ID,
                        range=f"I{row_idx + 1}:K{row_idx + 1}",
                        valueInputOption="RAW",
                        body={"values": [[exhibitor_count, large_booths, 0]]}
                    ).execute()
                    
                    results.append({
                        "show_name": show_name,
                        "status": "completed",
                        "exhibitor_count": exhibitor_count,
                        "file_link": drive_link
                    })
                    
                finally:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
        
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = f"Processed {len(results)} requests"
        jobs[job_id]["results"] = results
        
    except Exception as e:
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["error"] = str(e)
        import traceback
        traceback.print_exc()


# ============================================================================
# API Endpoints
# ============================================================================

@app.get("/api/pending", response_model=List[PendingRequest])
async def get_pending_requests():
    """List pending requests from Google Sheet."""
    
    try:
        sheets_service = get_sheets_service()
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range="A:L"
        ).execute()
        
        rows = result.get("values", [])
        if len(rows) < 2:
            return []
        
        # Parse header
        headers = rows[0]
        header_map = {}
        for i, h in enumerate(headers):
            key = h.lower().replace(" ", "_") if h else f"col_{i}"
            header_map[key] = i
        
        pending = []
        for idx, row in enumerate(rows[1:], start=1):
            # Helper to safely get row value
            def get_row_val(key: str, default=None):
                col_idx = header_map.get(key, -1)
                if col_idx >= 0 and col_idx < len(row):
                    val = row[col_idx]
                    return val if val else default
                return default
            
            # Check if delivered is blank
            delivered = get_row_val("delivered")
            
            if not delivered:  # Pending
                try:
                    exhibitors_val = get_row_val("exhibitors")
                    exhibitors = int(exhibitors_val) if exhibitors_val and exhibitors_val.isdigit() else None
                except:
                    exhibitors = None
                
                try:
                    large_booths_val = get_row_val("20x20+")
                    large_booths = int(large_booths_val) if large_booths_val and large_booths_val.isdigit() else None
                except:
                    large_booths = None
                
                pending.append(PendingRequest(
                    row_index=idx,
                    date_requested=get_row_val("date_requested"),
                    show_name=get_row_val("show_name"),
                    start_date=get_row_val("start"),
                    end_date=get_row_val("end"),
                    location=get_row_val("location"),
                    floorplan_url=get_row_val("floorplan"),
                    exhibitor_list_url=get_row_val("exhibitor_list"),
                    delivered=delivered,
                    exhibitors=exhibitors,
                    large_booths=large_booths,
                    missing_contact_info=get_row_val("missing_contact_info")
                ))
        
        return pending
        
    except Exception as e:
        print(f"Error fetching pending requests: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/process")
async def process_requests(request: ProcessRequest, background_tasks: BackgroundTasks):
    """Start processing selected requests."""
    
    if not request.row_indices:
        raise HTTPException(status_code=400, detail="No requests selected")
    
    # Get Seamless API key
    seamless_api_key = os.getenv("SEAMLESS_API_KEY", "")
    if not seamless_api_key:
        print("Warning: SEAMLESS_API_KEY not configured, enrichment will be skipped")
    
    # Create job
    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status": "pending",
        "progress": 0,
        "message": "Job created",
        "results": None,
        "error": None
    }
    
    # Start background processing
    background_tasks.add_task(process_request, job_id, request.row_indices, seamless_api_key)
    
    return {"job_id": job_id}


@app.get("/api/status/{job_id}", response_model=JobStatus)
async def get_job_status(job_id: str):
    """Check processing status."""
    
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        progress=job["progress"],
        message=job["message"],
        results=job.get("results"),
        error=job.get("error")
    )


@app.get("/api/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
